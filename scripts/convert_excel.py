#!/usr/bin/env python3
"""Convert the master Excel workbook to browser-ready JSON without geocoding."""

from __future__ import annotations

import argparse
import json
import math
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "Master Facilities v2"
REQUIRED_COLUMNS = [
    "Vertical", "Ultimate Parent", "Operating / Bottler Entity", "Facility Name",
    "Facility Type", "Country", "Latitude", "Longitude"
]


def clean(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, str):
        value = re.sub(r"\s+", " ", value).strip()
        return value or None
    return value


def coordinate(value, low, high):
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) and low <= result <= high else None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path, nargs="?", default=Path("data/facilities.json"))
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Missing required sheet: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    coordinate_audit = {}
    if "Coordinate Audit" in workbook.sheetnames:
        audit_rows = workbook["Coordinate Audit"].iter_rows(values_only=True)
        audit_headers = [clean(value) for value in next(audit_rows)]
        for audit_values in audit_rows:
            audit_item = dict(zip(audit_headers, audit_values))
            if audit_item.get("Excel Row") is not None:
                coordinate_audit[int(audit_item["Excel Row"])] = audit_item
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise SystemExit(f"Missing required columns: {', '.join(missing)}")

    facilities = []
    for excel_row, values in enumerate(rows, start=2):
        raw = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if not any(value is not None for value in raw.values()):
            continue
        lat = coordinate(raw.get("Latitude"), -90, 90)
        lon = coordinate(raw.get("Longitude"), -180, 180)
        mapped = lat is not None and lon is not None
        audit = coordinate_audit.get(excel_row, {})
        facilities.append({
            "id": f"facility-{excel_row}",
            "excelRow": excel_row,
            "vertical": raw.get("Vertical"),
            "ultimateParent": raw.get("Ultimate Parent"),
            "operatingEntity": raw.get("Operating / Bottler Entity"),
            "facilityName": raw.get("Facility Name"),
            "facilityType": raw.get("Facility Type"),
            "streetAddress": raw.get("Street Address"),
            "city": raw.get("City"),
            "stateProvince": raw.get("State / Province"),
            "postalCode": str(raw["Postal Code"]) if raw.get("Postal Code") is not None else None,
            "country": raw.get("Country"),
            "latitude": lat if mapped else None,
            "longitude": lon if mapped else None,
            "mapped": mapped,
            "ace": raw.get("ACE"),
            "verificationStatus": raw.get("Verification Status"),
            "sourceUrl": raw.get("Source URL"),
            "researchNotes": raw.get("Research Notes"),
            "coordinateStatus": audit.get("Coordinate Status"),
            "coordinateMethod": audit.get("Method / Source"),
            "coordinateQuality": audit.get("Match Quality"),
            "coordinateNote": audit.get("Matched Address / Note"),
        })

    payload = {
        "metadata": {
            "sourceFile": args.input.name,
            "sourceSheet": SHEET_NAME,
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "totalRecords": len(facilities),
            "mappedRecords": sum(item["mapped"] for item in facilities),
            "unmappedRecords": sum(not item["mapped"] for item in facilities),
        },
        "facilities": facilities,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"Converted {len(facilities)} records: {payload['metadata']['mappedRecords']} mapped, {payload['metadata']['unmappedRecords']} unmapped")


if __name__ == "__main__":
    main()
